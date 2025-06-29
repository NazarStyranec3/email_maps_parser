
import os
from openpyxl import Workbook, load_workbook
def add_ex(row_data):
    filename = '../m.xlsx'

    headers = ['силка на карту googl', 'силка на сайт', 'назва',
               'номер з карти googl', 'пошта силка на сайт', 'номер силка на сайт']

    # Перетворення списків у рядки з комами
    clean_row = [", ".join(item) if isinstance(item, list) else item for item in row_data]

    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        ws.append(clean_row)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append(clean_row)

    wb.save(filename)
    print(f"Дані збережено в '{filename}'")